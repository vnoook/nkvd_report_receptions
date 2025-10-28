# "Врач, оказавший услугу", "Услуга", "Отделение сотрудника", "Вид оплаты"
import openpyxl
import pandas as pd

# файл данных
file_xlsx = 'Аналитика2609-2710.xlsx'

# открывается файл
wb = openpyxl.load_workbook(file_xlsx)
wb_s = wb.active

# строки начала и конца
wb_file_sheet_row_begin = 3
wb_file_sheet_row_end = wb_s.max_row - 1
wb_file_sheet_col_begin = wb_s.min_column
wb_file_sheet_col_end = wb_s.max_column

# получение всех данных из файла и его закрытие
list_all_data = []
for row in range(wb_file_sheet_row_begin, wb_file_sheet_row_end + 1):
    otdel = wb_s.cell(row=row, column=10).value
    doc = wb_s.cell(row=row, column=2).value
    service = wb_s.cell(row=row, column=3).value
    payment = wb_s.cell(row=row, column=22).value
    list_all_data.append([otdel, doc, service, payment])
wb.close()
# print(*list_all_data_fresh, sep='\n')

# подсчёт данных
otdel_dict = {}
for row in list_all_data:
    otdel = row[0]
    if otdel_dict.get(otdel) is None:
        otdel_dict[otdel] = 1
    else:
        otdel_dict[otdel] = otdel_dict[otdel] + 1
# print(*otdel_dict, sep='\n')

doc_dict = {}
for otdel in otdel_dict.keys():
    # print(otdel)
    for row in list_all_data:
        if otdel == row[0]:
            # print(row[1])
            pass
    # print('************')

# //////////////////////////////////////////////////////////////
# Создание DataFrame на основе
df0 = pd.DataFrame(list_all_data, columns=['otdel', 'doc', 'service', 'payment'])
print(df0)
df0.to_excel('out0.xlsx', sheet_name='Sheet1')
print('*'*55)

# df1 = df0.groupby('otdel').count()
# print(df1)
# df1.to_excel('out1.xlsx', sheet_name='Sheet1')
# print('*'*55)
# df2 = df0.groupby('doc').count()
# print(df2)
# df2.to_excel('out2.xlsx', sheet_name='Sheet1')
# print('*'*55)
# df3 = df0.groupby('service').count()
# print(df3)
# df3.to_excel('out3.xlsx', sheet_name='Sheet1')
# print('*'*55)
# df4 = df0.groupby(['otdel', 'doc', 'service', 'payment']).count()
# print(df4)
# df4.to_excel('out4.xlsx', sheet_name='Sheet1')
# print('*'*55)

df5 = pd.crosstab(df0['otdel'], df0['doc'], margins=True)
print(df5)
df5.to_excel('out5.xlsx', sheet_name='Sheet1')
print('*'*55)
exit()

q_prod_name = df.pivot_table('payment', ['otdel','doc','service','payment'], aggfunc='count', fill_value = 0)
q_prod_name.to_excel('out20.xlsx', sheet_name='Sheet1')




# q_prod_name = df.pivot_table(values=[],
#               index=['otdel','doc','service','payment'],
#               columns=['otdel','doc','service','payment'],
#               aggfunc='count')
# q_prod_name.to_excel('out20.xlsx', sheet_name='Sheet1')
# q_prod_name = df.pivot_table(values=['otdel','doc','service','payment'],
#               index=[],
#               columns=['otdel','doc','service','payment'],
#               aggfunc='count')
# q_prod_name.to_excel('out21.xlsx', sheet_name='Sheet1')
# q_prod_name = df.pivot_table(values=['otdel','doc','service','payment'],
#               index=['otdel','doc','service','payment'],
#               columns=[],
#               aggfunc='count')
# q_prod_name.to_excel('out22.xlsx', sheet_name='Sheet1')
# q_prod_name = df.pivot_table(values=['otdel','doc','service','payment'],
#               index=['otdel','doc','service','payment'],
#               columns=['otdel','doc','service','payment'])
# q_prod_name.to_excel('out23.xlsx', sheet_name='Sheet1')

# df_group1 = df_group1.reset_index()
# for index, row in df_group1.iterrows():
#     for val in ['doc', 'service', 'otdel', 'money']:
#         print(f'{row[val] = }')
#     print('*'*155)
# df_group1.to_excel('out3.xlsx')
