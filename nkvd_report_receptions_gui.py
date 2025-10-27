import os
import sys
import openpyxl
import openpyxl.utils
import openpyxl.styles
import PyQt5
import PyQt5.QtWidgets
import PyQt5.QtCore
import PyQt5.QtGui


# класс главного окна
class WindowMain(PyQt5.QtWidgets.QMainWindow):
    """Класс главного окна"""

    # описание главного окна
    def __init__(self):
        super().__init__()

        # переменные
        self.info_extention_open_file_xlsx = 'Файлы XLSX (*.xlsx)'
        self.info_path_open_file = None
        self.text_empty_path_file = 'файл пока не выбран'
        self.info_for_open_file = 'Выберите XLSX файл (.XLSX)'

        # главное окно, надпись на нём и размеры
        self.setWindowTitle('Парсер XLSX файлов для отчёта Журнал записей пациентов')
        self.setGeometry(450, 100, 700, 350)
        self.setWindowFlags(PyQt5.QtCore.Qt.WindowStaysOnTopHint)

        # ОБЪЕКТЫ НА ФОРМЕ
        # выбор свежего файла
        # label_select_file_fresh_data
        self.label_select_file_fresh_data = PyQt5.QtWidgets.QLabel(self)
        self.label_select_file_fresh_data.setObjectName('label_select_file_fresh_data')
        self.label_select_file_fresh_data.setText('1) Выберите свежий файл XLSX')
        self.label_select_file_fresh_data.setGeometry(PyQt5.QtCore.QRect(10, 10, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_select_file_fresh_data.setFont(font)
        self.label_select_file_fresh_data.adjustSize()
        self.label_select_file_fresh_data.setToolTip(self.label_select_file_fresh_data.objectName())

        # toolButton_select_fresh_xlsx
        self.toolButton_select_fresh_xlsx = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_fresh_xlsx.setObjectName('toolButton_select_fresh_xlsx')
        self.toolButton_select_fresh_xlsx.setText('...')
        self.toolButton_select_fresh_xlsx.setGeometry(PyQt5.QtCore.QRect(10, 40, 50, 20))
        self.toolButton_select_fresh_xlsx.setFixedWidth(50)
        self.toolButton_select_fresh_xlsx.clicked.connect(self.select_file_fresh_xlsx)
        self.toolButton_select_fresh_xlsx.setToolTip(self.toolButton_select_fresh_xlsx.objectName())

        # label_path_fresh_file
        self.label_path_fresh_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_fresh_file.setObjectName('label_path_fresh_file')
        self.label_path_fresh_file.setEnabled(False)
        self.label_path_fresh_file.setText(self.text_empty_path_file)
        self.label_path_fresh_file.setGeometry(PyQt5.QtCore.QRect(10, 70, 400, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_path_fresh_file.setFont(font)
        self.label_path_fresh_file.adjustSize()
        self.label_path_fresh_file.setToolTip(self.label_path_fresh_file.objectName())

        # выбор старого файла
        # label_select_file_old_data
        self.label_select_file_old_data = PyQt5.QtWidgets.QLabel(self)
        self.label_select_file_old_data.setObjectName('label_select_file_old_data')
        self.label_select_file_old_data.setText('2) Выберите старый файл XLSX')
        self.label_select_file_old_data.setGeometry(PyQt5.QtCore.QRect(10, 100, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_select_file_old_data.setFont(font)
        self.label_select_file_old_data.adjustSize()
        self.label_select_file_old_data.setToolTip(self.label_select_file_old_data.objectName())

        # toolButton_select_old_xlsx
        self.toolButton_select_old_xlsx = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_old_xlsx.setObjectName('toolButton_select_old_xlsx')
        self.toolButton_select_old_xlsx.setText('...')
        self.toolButton_select_old_xlsx.setGeometry(PyQt5.QtCore.QRect(10, 130, 50, 20))
        self.toolButton_select_old_xlsx.setFixedWidth(50)
        self.toolButton_select_old_xlsx.clicked.connect(self.select_file_old_xlsx)
        self.toolButton_select_old_xlsx.setToolTip(self.toolButton_select_old_xlsx.objectName())

        # label_path_old_file
        self.label_path_old_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_old_file.setObjectName('label_path_old_file')
        self.label_path_old_file.setEnabled(False)
        self.label_path_old_file.setText(self.text_empty_path_file)
        self.label_path_old_file.setGeometry(PyQt5.QtCore.QRect(10, 160, 400, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_path_old_file.setFont(font)
        self.label_path_old_file.adjustSize()
        self.label_path_old_file.setToolTip(self.label_path_old_file.objectName())

        # checkBox_short
        self.checkBox_short = PyQt5.QtWidgets.QCheckBox(self)
        self.checkBox_short.setObjectName('checkBox_short')
        self.checkBox_short.setGeometry(PyQt5.QtCore.QRect(10, 190, 200, 40))
        self.checkBox_short.clicked.connect(self.checkbox_click)
        self.checkBox_short.setChecked(True)
        self.checkBox_short.setText('Хочу короткий отчёт')
        self.checkBox_short.setToolTip(self.checkBox_short.objectName())

        # pushButton_parse_to_xls
        self.pushButton_parse_to_xls = PyQt5.QtWidgets.QPushButton(self)
        self.pushButton_parse_to_xls.setObjectName('pushButton_parse_to_xls')
        self.pushButton_parse_to_xls.setEnabled(False)
        self.pushButton_parse_to_xls.setText('Создать отчёт "Журнал записей пациентов"')
        self.pushButton_parse_to_xls.setGeometry(PyQt5.QtCore.QRect(10, 230, 260, 25))
        self.pushButton_parse_to_xls.clicked.connect(self.parse_xlsx)
        self.pushButton_parse_to_xls.setToolTip(self.pushButton_parse_to_xls.objectName())

        # EXIT
        # button_exit
        self.button_exit = PyQt5.QtWidgets.QPushButton(self)
        self.button_exit.setObjectName('button_exit')
        self.button_exit.setText('Выход')
        self.button_exit.setGeometry(PyQt5.QtCore.QRect(10, 300, 100, 25))
        # self.button_exit.setFixedWidth(50)
        self.button_exit.clicked.connect(self.click_on_btn_exit)
        self.button_exit.setToolTip(self.button_exit.objectName())

    # событие - нажатие на чекбокс выбора размера отчёта
    def checkbox_click(self):
        # print(self.checkBox_short.isChecked())
        pass

    # событие - нажатие на кнопку выбора файла
    def select_file_fresh_xlsx(self):
        # переменная для хранения информации из окна выбора файла
        data_of_open_file_name = None

        # запоминание старого значения пути выбора файлов
        old_path_of_selected_xlsx_file = self.label_path_fresh_file.text()

        # непосредственное окно выбора файла и переменная для хранения пути файла
        data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                             self.info_for_open_file,
                                                                             self.info_path_open_file,
                                                                             self.info_extention_open_file_xlsx)

        # выбор только пути файла из data_of_open_file_name
        file_name = data_of_open_file_name[0]

        # нажата кнопка выбора XLSX файла
        if file_name == '':
            self.label_path_fresh_file.setText(old_path_of_selected_xlsx_file)
            self.label_path_fresh_file.adjustSize()
        else:
            old_path_of_selected_xlsx_file = self.label_path_fresh_file.text()
            self.label_path_fresh_file.setText(file_name)
            self.label_path_fresh_file.adjustSize()

        # активация и деактивация объектов
        self.units_activate()

    # событие - нажатие на кнопку выбора файла
    def select_file_old_xlsx(self):
        # переменная для хранения информации из окна выбора файла
        data_of_open_file_name = None

        # запоминание старого значения пути выбора файлов
        old_path_of_selected_xlsx_file = self.label_path_old_file.text()

        # непосредственное окно выбора файла и переменная для хранения пути файла
        data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                             self.info_for_open_file,
                                                                             self.info_path_open_file,
                                                                             self.info_extention_open_file_xlsx)

        # выбор только пути файла из data_of_open_file_name
        file_name = data_of_open_file_name[0]

        # нажата кнопка выбора XLSX файла
        if file_name == '':
            self.label_path_old_file.setText(old_path_of_selected_xlsx_file)
            self.label_path_old_file.adjustSize()
        else:
            old_path_of_selected_xlsx_file = self.label_path_old_file.text()
            self.label_path_old_file.setText(file_name)
            self.label_path_old_file.adjustSize()

        # активация и деактивация объектов
        self.units_activate()

    # активация и деактивация объектов на форме зависящее от выбора файла
    def units_activate(self):
        # активация и деактивация объектов на форме зависящее от выбора файла
        if ((self.text_empty_path_file not in self.label_path_fresh_file.text()) and
                (self.text_empty_path_file not in self.label_path_old_file.text())):
            self.pushButton_parse_to_xls.setEnabled(True)

    # функция создания отчёта
    def parse_xlsx(self):
        # получение путей и имён выбранных файлов
        file_xlsx_fresh = self.label_path_fresh_file.text()
        file_xlsx_path_fresh = os.path.split(file_xlsx_fresh)[0]
        file_xlsx_name_fresh = os.path.split(file_xlsx_fresh)[1]

        file_xlsx_old = self.label_path_old_file.text()
        file_xlsx_path_old = os.path.split(file_xlsx_old)[0]
        file_xlsx_name_old = os.path.split(file_xlsx_old)[1]

        # структуры для сбора данных
        list_all_data_fresh = []
        list_all_data_old = []

        # словарь для хранения "отделений"
        dict_departments_fresh = {}
        dict_departments_old = {}
        # словарь для хранения "записавших организаций"
        dict_organization_fresh = {}
        dict_organization_old = {}
        # словарь для хранения "кем записан"
        dict_persona_fresh = {}
        dict_persona_old = {}
        # словарь для хранения "статуса услуги"
        dict_status_service_fresh = {}
        dict_status_service_old = {}

        # персоны, которые нужно суммировать
        str_person_summ = {
            'Интеграция Е.Р.': 'ЕПГУ-Госуслуги',
            'Система Г.С.': 'СГС-Робот Николай',
            'Administrator A.A.': 'МИАЦ'
        }
        # строки по которым нужно фильтровать
        str_for_filter = (
            'Амбулаторное отделение №1',
            'Амбулаторное отделение №2',
            'Амбулаторное отделение №3',
            'Амбулаторное отделение №4',
            'Подростковый специализированный центр профилактики и лечения инфекций, передаваемых половым путем'
        )

        # номера колонок для сбора данных
        wb_in_s_col_1 = 7  # Отделение
        wb_in_s_col_2 = 18  # Записавшая организация
        wb_in_s_col_3 = 19  # Кем записан
        wb_in_s_col_4 = 9  # Услуга
        wb_in_s_col_5 = 10  # Статус услуги

        # открывается свежий файл
        wb_in = openpyxl.load_workbook(file_xlsx_fresh)
        wb_in_s = wb_in['Журнал записей пациентов']

        # строка начала и конца
        wb_in_s_row_begin = 3
        wb_in_s_row_end = wb_in_s.max_row - 1

        # получение всех данных из файла и его закрытие, чтобы к нему больше не возвращаться
        for row in range(wb_in_s_row_begin, wb_in_s_row_end + 1):
            list_all_data_fresh.append([wb_in_s.cell(row=row, column=wb_in_s_col_1).value,
                                        wb_in_s.cell(row=row, column=wb_in_s_col_2).value,
                                        wb_in_s.cell(row=row, column=wb_in_s_col_3).value,
                                        wb_in_s.cell(row=row, column=wb_in_s_col_4).value,
                                        wb_in_s.cell(row=row, column=wb_in_s_col_5).value
                                        ])
        wb_in.close()

        # открывается старый файл
        wb_in = openpyxl.load_workbook(file_xlsx_old)
        wb_in_s = wb_in['Журнал записей пациентов']

        # строка начала и конца
        wb_in_s_row_begin = 3
        wb_in_s_row_end = wb_in_s.max_row - 1

        # получение всех данных из файла и его закрытие, чтобы к нему больше не возвращаться
        for row in range(wb_in_s_row_begin, wb_in_s_row_end + 1):
            list_all_data_old.append([wb_in_s.cell(row=row, column=wb_in_s_col_1).value,
                                      wb_in_s.cell(row=row, column=wb_in_s_col_2).value,
                                      wb_in_s.cell(row=row, column=wb_in_s_col_3).value,
                                      wb_in_s.cell(row=row, column=wb_in_s_col_4).value,
                                      wb_in_s.cell(row=row, column=wb_in_s_col_5).value
                                      ])
        wb_in.close()

        # создание нового отчёта в xlsx и активация рабочего листа
        wb_out = openpyxl.Workbook()
        wb_out_s = wb_out.active

        # подсчёт и распределение свежих данных
        for val_str in list_all_data_fresh:
            if val_str[0] in str_for_filter:
                # заполнение словаря отделений
                if dict_departments_fresh.get(val_str[0]) is None:
                    dict_departments_fresh[val_str[0]] = 1
                else:
                    dict_departments_fresh[val_str[0]] = dict_departments_fresh[val_str[0]] + 1

                # заполнение словаря записавших организаций
                if dict_organization_fresh.get(val_str[0]) is None:
                    dict_organization_fresh[val_str[0]] = {val_str[1]: 1}
                else:
                    if val_str[1] is not None:
                        if dict_organization_fresh[val_str[0]].get(val_str[1]) is None:
                            dict_organization_fresh[val_str[0]][val_str[1]] = 1
                        else:
                            dict_organization_fresh[val_str[0]][val_str[1]] = \
                                dict_organization_fresh[val_str[0]][val_str[1]] + 1

                # заполнение словаря "кем записан"
                if val_str[2] in str_person_summ:
                    if dict_persona_fresh.get(val_str[0]) is None:
                        dict_persona_fresh[val_str[0]] = {val_str[2]: 1}
                    else:
                        if val_str[2] is not None:
                            if dict_persona_fresh[val_str[0]].get(val_str[2]) is None:
                                dict_persona_fresh[val_str[0]][val_str[2]] = 1
                            else:
                                dict_persona_fresh[val_str[0]][val_str[2]] = \
                                    dict_persona_fresh[val_str[0]][val_str[2]] + 1

                # заполнение словаря "статус услуги"
                if dict_status_service_fresh.get(val_str[0]) is None:
                    dict_status_service_fresh[val_str[0]] = {val_str[4]: 1}
                else:
                    if val_str[4] is not None:
                        if dict_status_service_fresh[val_str[0]].get(val_str[4]) is None:
                            dict_status_service_fresh[val_str[0]][val_str[4]] = 1
                        else:
                            dict_status_service_fresh[val_str[0]][val_str[4]] = \
                                dict_status_service_fresh[val_str[0]][val_str[4]] + 1

        # подсчёт и распределение старых данных
        for val_str in list_all_data_old:
            if val_str[0] in str_for_filter:
                # заполнение словаря отделений
                if dict_departments_old.get(val_str[0]) is None:
                    dict_departments_old[val_str[0]] = 1
                else:
                    dict_departments_old[val_str[0]] = dict_departments_old[val_str[0]] + 1

                # заполнение словаря записавших организаций
                if dict_organization_old.get(val_str[0]) is None:
                    dict_organization_old[val_str[0]] = {val_str[1]: 1}
                else:
                    if val_str[1] is not None:
                        if dict_organization_old[val_str[0]].get(val_str[1]) is None:
                            dict_organization_old[val_str[0]][val_str[1]] = 1
                        else:
                            dict_organization_old[val_str[0]][val_str[1]] = \
                                dict_organization_old[val_str[0]][val_str[1]] + 1

                # заполнение словаря "кем записан"
                if val_str[2] in str_person_summ:
                    if dict_persona_old.get(val_str[0]) is None:
                        dict_persona_old[val_str[0]] = {val_str[2]: 1}
                    else:
                        if val_str[2] is not None:
                            if dict_persona_old[val_str[0]].get(val_str[2]) is None:
                                dict_persona_old[val_str[0]][val_str[2]] = 1
                            else:
                                dict_persona_old[val_str[0]][val_str[2]] = \
                                    dict_persona_old[val_str[0]][val_str[2]] + 1

                # заполнение словаря "статус услуги"
                if dict_status_service_old.get(val_str[0]) is None:
                    dict_status_service_old[val_str[0]] = {val_str[4]: 1}
                else:
                    if val_str[4] is not None:
                        if dict_status_service_old[val_str[0]].get(val_str[4]) is None:
                            dict_status_service_old[val_str[0]][val_str[4]] = 1
                        else:
                            dict_status_service_old[val_str[0]][val_str[4]] = \
                                dict_status_service_old[val_str[0]][val_str[4]] + 1

        # сортировка словарей свежих
        dict_organization_fresh = dict(sorted(dict_organization_fresh.items()))
        dict_departments_fresh = dict(sorted(dict_departments_fresh.items()))
        dict_persona_fresh = dict(sorted(dict_persona_fresh.items()))
        dict_status_service_fresh = dict(sorted(dict_status_service_fresh.items()))

        # сортировка словарей старых
        dict_organization_old = dict(sorted(dict_organization_old.items()))
        dict_departments_old = dict(sorted(dict_departments_old.items()))
        dict_persona_old = dict(sorted(dict_persona_old.items()))
        dict_status_service_old = dict(sorted(dict_status_service_old.items()))

        # добавления стиля строк
        style0 = openpyxl.styles.Font(bold=True, size=20)
        style1 = openpyxl.styles.Font(bold=True, size=14)
        style2 = openpyxl.styles.Font(bold=True, size=12)
        style3 = openpyxl.styles.Font(bold=False, size=12)
        style4 = openpyxl.styles.Font(italic=True, size=12)

        row = 1
        col = 1
        persona_string = ''

        # установка ширины колонки
        wb_out_s.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 85

        # добавление шапки
        wb_out_s.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
        wb_out_s.cell(row=row, column=col).font = style0
        wb_out_s.cell(row=row, column=col).value = f'Журнал записей пациентов'
        row += 1
        wb_out_s.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
        wb_out_s.cell(row=row, column=col).font = style1
        wb_out_s.cell(row=row, column=col).value = f'(данные за неделю __)'
        row += 1
        wb_out_s.cell(row=row, column=col).font = style3
        wb_out_s.cell(row=row, column=col).value = f''
        row += 1

        # формирование отчёта
        for k_org, v_org in dict_organization_fresh.items():
            # добавление первой строки
            wb_out_s.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
            wb_out_s.cell(row=row, column=col).font = style1
            wb_out_s.cell(row=row, column=col).value = f'{k_org} - {dict_departments_fresh[k_org]} пациент(ов)'
            row += 1

            # добавление второй строки
            if dict_persona_fresh[k_org]:
                persona_string = ''
                for str_person in str_person_summ:
                    if dict_persona_fresh[k_org].get(str_person):
                        persona_string = persona_string + (f'{dict_persona_fresh[k_org].get(str_person)}'
                                                           f' через {str_person_summ[str_person]}') + ', '

            wb_out_s.cell(row=row, column=col).font = style2
            wb_out_s.cell(row=row, column=col).value = f'(из них {persona_string[:-2]})'
            row += 1

            # добавление строк по организациям
            if self.checkBox_short.isChecked():
                # короткий отчёт
                q_sum = 0
                for d, q in v_org.items():
                    if d == 'ГБУЗ НСО «НОККВД»':
                        wb_out_s.cell(row=row, column=col).font = style3
                        wb_out_s.cell(row=row, column=col).value = f'{d} - {q}'
                        row += 1
                    else:
                        q_sum = q_sum + int(q)
                wb_out_s.cell(row=row, column=col).font = style3
                wb_out_s.cell(row=row, column=col).value = f'Другие МО - {q_sum}'
                row += 1
            else:
                # длинный отчёт
                for d, q in v_org.items():
                    wb_out_s.cell(row=row, column=col).font = style3
                    wb_out_s.cell(row=row, column=col).value = f'{d} - {q}'
                    row += 1

            # добавление строки про "статус услуги"
            if dict_status_service_fresh[k_org]:
                status_string = ''
                for k_p, v_p in dict_status_service_fresh[k_org].items():
                    status_string = status_string + f'{k_p} - {v_p}' + ', '

                wb_out_s.cell(row=row, column=col).font = style4
                wb_out_s.cell(row=row, column=col).value = status_string[:-2]
                row += 1

            # добавление пустой строки разделения между "отделениями"
            wb_out_s.append([])

            # формирование предыдущего периода
            # добавление первой строки
            thin_border = openpyxl.styles.borders.Border(top=openpyxl.styles.borders.Side(style='thin'))
            wb_out_s.cell(row=row, column=col).border = thin_border
            wb_out_s.cell(row=row, column=col).font = style3
            wb_out_s.cell(row=row, column=col).value = f'Предыдущая неделя - {dict_departments_old[k_org]} пациент(ов)'
            row += 1

            # добавление второй строки
            if dict_persona_old[k_org]:
                persona_string = ''
                for str_person in str_person_summ:
                    if dict_persona_old[k_org].get(str_person):  # is not None
                        persona_string = persona_string + (f'{dict_persona_old[k_org].get(str_person)}'
                                                           f' через {str_person_summ[str_person]}') + ', '

            wb_out_s.cell(row=row, column=col).font = style3
            wb_out_s.cell(row=row, column=col).value = f'(из них {persona_string[:-2]})'
            row += 1

            # добавление пустой строки разделения между "отделениями"
            wb_out_s.append([])
            wb_out_s.cell(row=row, column=col).font = style3
            wb_out_s.cell(row=row, column=col).value = f''
            row += 1

        # создание названия выходного файла xls
        file_xls_path = file_xlsx_path_fresh[:]
        file_xls_name = os.path.splitext(file_xlsx_name_fresh)[0] + '_отчёт.xlsx'
        file_report = os.path.abspath(os.path.join(file_xls_path, file_xls_name))

        # сохранение файла xlsx и закрытие его
        wb_out.save(file_report)
        wb_out.close()

        # открытие папки с сохранённым файлом xls
        fullpath = os.path.abspath(file_xls_path)
        PyQt5.QtGui.QDesktopServices.openUrl(PyQt5.QtCore.QUrl.fromLocalFile(fullpath))

    # событие - нажатие на кнопку Выход
    @staticmethod
    def click_on_btn_exit():
        sys.exit()


# создание основного окна
def main_app():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    app_window_main = WindowMain()
    app_window_main.show()
    sys.exit(app.exec_())


# запуск основного окна
if __name__ == '__main__':
    main_app()
